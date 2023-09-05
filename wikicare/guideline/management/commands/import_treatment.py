#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 14:19:08 2023

@author: anagoncalves
"""

import openpyxl
from django.core.management.base import BaseCommand
from wikicare.models import Treatment, Symptom

class Command(BaseCommand):
    help = 'Import treatments from XLSX file'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_file', type=str, help='Path to the XLSX file')

    def handle(self, *args, **options):
        xlsx_file = options['xlsx_file']

        # Open the XLSX file
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active

        symptoms = []
        treatments = []

        # Read symptoms and treatments from the XLSX file
        for row in sheet.iter_rows(values_only=True):
            if row[0]:  # Check if a symptom is present
                symptoms.append(row[0])
            if row[1:]:  # Check if treatments are present
                treatments.extend(row[1:])

        # Remove any empty values or duplicates
        symptoms = list(filter(None, symptoms))
        treatments = list(filter(None, treatments))
        treatments = list(set(treatments))

        # Create symptoms and treatments
        for symptom in symptoms:
            symptom_obj, _ = Symptom.objects.get_or_create(name=symptom)
            for treatment in treatments:
                if sheet.cell(row=symptoms.index(symptom) + 1, column=treatments.index(treatment) + 2).value == 'x':
                    Treatment.objects.create(symptom=symptom_obj, name=treatment)

        self.stdout.write(self.style.SUCCESS('Successfully imported treatments from XLSX file.'))
        
        
        
        
        
        
        